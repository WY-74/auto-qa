import openpyxl
import traceback
from openpyxl.styles import PatternFill, Font
from base.web import WebKeys
from config import log


class ExcelDriver:
    def __init__(self) -> None:
        self.logger = log.get_logger()

    def get_parameter(self, raw: str):
        parameters = dict()
        if raw:
            for single in raw.split(";"):
                key, value = single.split("=", 1)
                parameters[key] = value

        return parameters

    def run(self, data_path):
        excel = openpyxl.load_workbook(data_path)

        for sheet in excel.sheetnames:
            if sheet.startswith("#"):  # 跳过某些用例
                continue

            web = WebKeys()
            for column in excel[sheet].values:
                if isinstance(column[0], int):
                    self.logger.info(column[3])
                    try:
                        parameters = self.get_parameter(column[2])
                        getattr(web, column[1])(**parameters)
                        self.set_to_pass(excel[sheet].cell(row=column[0] + 1, column=5))
                    except:
                        traceback.print_exc()
                        self.set_to_failed(excel[sheet].cell(row=column[0] + 1, column=5))
            web.close()

        excel.save(data_path)

    def set_to_failed(self, cell):
        cell.value = 'FAILED'
        cell.fill = PatternFill("solid", "FF0000")
        cell.font = Font(bold=True)

    def set_to_pass(self, cell):
        cell.value = 'PASS'
        cell.fill = PatternFill("solid", "AACF91")
        cell.font = Font(bold=True)
